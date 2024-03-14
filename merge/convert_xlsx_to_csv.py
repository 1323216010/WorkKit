from openpyxl import load_workbook
import csv
import os

folder_path = r'C:\Users\pengcheng.yan\Desktop\test_pit'
output_path = r'C:\Users\pengcheng.yan\Desktop\test_pit'

if not os.path.exists(output_path):
    os.makedirs(output_path)

for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active  # 或者使用 wb.get_sheet_by_name('Sheet1') 来选择特定的工作表

        output_file_path = os.path.join(output_path, f"{os.path.splitext(filename)[0]}.csv")

        with open(output_file_path, 'w', newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
